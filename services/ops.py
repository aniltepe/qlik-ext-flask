import os
from dotenv import load_dotenv
import csv
import openpyxl
from datetime import datetime
import services.db as db
from services.meta import FieldType, new_uuid, format_py_replace, db_field_type
from services.qlikprops import *
from services.qlikengine import QlikEngine

load_dotenv()

QLIK_WS_URL = os.getenv("QLIK_WS_URL")
QLIK_WS_CERT_PATH = os.getenv("QLIK_WS_CERT_PATH")
QLIK_WS_USER = os.getenv("QLIK_WS_USER")
QLIK_META_APP = os.getenv("QLIK_META_APP")
DB_CONN_META = os.getenv("DB_CONN_META")
CLIENT_URL = os.getenv("CLIENT_URL")

# print(QLIK_WS_URL, QLIK_WS_CERT_PATH, QLIK_WS_USER, QLIK_META_APP, DB_CONN_META, CLIENT_URL)

TABLE_PREFIX = "QDE_"
QLIK_TAB_PREFIX = "///$tab "
QLIK_SECTION = "QlikDataEntry"

qlik = QlikEngine(url=QLIK_WS_URL, cert_path=QLIK_WS_CERT_PATH, user=QLIK_WS_USER)

def get_columns(appid, dbtable):
    fields = db.select(
        DB_CONN_META, 
        ["*"], 
        [("appid", "=", appid), "AND", ("dbtable", "=", dbtable)], 
        ["fieldorder"]
    )
    # print(fields)
    return fields

# konfigürasyon esnasında BI dev tarafından girilen field tanımı için yazılan dbtable ile ilişkili başka bir field yoksa yani dbtable ilk kez yaratılıyorsa
# o dbtable ile ilişkili 3 buton (ekle, düzenle, sil) ve datatable sheet'le birlikte yaratılır
def custom_sheet_create(appid, title, fields):
    table_id, mbox_add_id, mbox_edit_id, mbox_del_id = new_uuid(), new_uuid(), new_uuid(), new_uuid()
    btn_cancel_id, btn_add_id, btn_edit_id, btn_del_id = new_uuid(), new_uuid(), new_uuid(), new_uuid()
    sheet_id = qlik.create_object(appid, create_sheet_prop_empty())
    container_table_id = qlik.create_child(appid, sheet_id, create_container_prop_table(table_id, mbox_add_id, mbox_edit_id, mbox_del_id))
    container_btn2_id = qlik.create_child(appid, sheet_id, create_container_prop_btn2(btn_add_id, btn_cancel_id, "EKLE", "İPTAL"))
    container_btn1_1_id = qlik.create_child(appid, sheet_id, create_container_prop_btn1(btn_edit_id, "DÜZENLE"))
    container_btn1_2_id = qlik.create_child(appid, sheet_id, create_container_prop_btn1(btn_del_id, "SİL"))
    table_qid = qlik.create_child(appid, container_table_id, create_table_prop(table_id, fields))
    qlik.create_child(appid, container_table_id, create_mediabox_prop(title, mbox_add_id, "add"))
    qlik.create_child(appid, container_table_id, create_mediabox_prop(title, mbox_edit_id, "edit"))
    qlik.create_child(appid, container_table_id, create_mediabox_prop(title, mbox_del_id, "delete"))
    qlik.create_child(appid, container_btn2_id, create_button_prop(btn_cancel_id, "0", "İPTAL"))
    qlik.create_child(appid, container_btn2_id, create_button_prop(btn_add_id, "1", "EKLE"))
    qlik.create_child(appid, container_btn1_1_id, create_button_prop(btn_edit_id, "2", "DÜZENLE"))
    qlik.create_child(appid, container_btn1_2_id, create_button_prop(btn_del_id, "3", "SİL"))
    qlik.set_properties(appid, sheet_id, create_sheet_prop(title, container_table_id, container_btn2_id, container_btn1_1_id, container_btn1_2_id))
    return (sheet_id, table_qid)

# konfigürasyon esnasında BI dev tarafından girilen dbtable değeriyle ilişkili mevcutta field tanımları ve dolayısıyla sheet objesi varsa
# o sheet'teki datatable'a yeni alan eklenir, sheet değişmeden datatable güncellenmiş olur (butonlarda da bir değişiklik olmaz)
def custom_sheet_update(appid, tableqid, fields):
    table_props = qlik.get_object(appid, tableqid)
    qlik.set_properties(appid, tableqid, create_table_prop(table_props["containerChildId"], fields, tableqid))

# son kullanıcının girdiği satırlar QDE_prefix'li ilgili tabloya insert edilir
def add_bulk(data, appid, dbtable):
    table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", appid), "AND", ("dbtable", "=", dbtable)])
    for row in data:
        date_adjusted = adjust_fields(row, table_fields)
        db.insert(f"[{TABLE_PREFIX}{dbtable}]", date_adjusted.keys(), date_adjusted.values())
    qlik.reload_app(appid)

# son kullanıcının güncellediği satırlar QDE_prefix'li ilgili tabloda update edilir
# tüm alanlarında tamamen aynı değere sahip birden fazla kayıt varsa top=1'den dolayı sadece birini günceller
# (tekrarlı kayıtların oluşmasını engelleyecek bir yapı bulunmuyor)
def edit_bulk(data, appid, dbtable):
    table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", appid), "AND", ("dbtable", "=", dbtable)])    
    for row in data:
        date_adjusted_old = adjust_fields(row["oldvalue"], table_fields)
        date_adjusted_new = adjust_fields(row["newvalue"], table_fields)
        where_condition = [(key, "=", value) if value is not None else f"{key} IS NULL" for key, value in date_adjusted_old.items()]
        i = 1
        while i < len(where_condition):
            where_condition.insert(i, "AND")
            i += 2
        set_condition = [(key, value) for key, value in date_adjusted_new.items()]
        db.update(f"[{TABLE_PREFIX}{dbtable}]", where_condition, set_condition, top=1)
    qlik.reload_app(appid)

# son kullanıcının güncellediği satırlar QDE_prefix'li ilgili tabloda delete edilir
# tüm alanlarında tamamen aynı değere sahip birden fazla kayıt varsa top=1'den dolayı sadece birini siler
# (tekrarlı kayıtların oluşmasını engelleyecek bir yapı bulunmuyor)
def delete_bulk(data, appid, dbtable):
    table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", appid), "AND", ("dbtable", "=", dbtable)])
    for row in data:
        date_adjusted_old = adjust_fields(row, table_fields)
        where_condition = [(key, "=", value) if value is not None else f"{key} IS NULL" for key, value in date_adjusted_old.items()]
        i = 1
        while i < len(where_condition):
            where_condition.insert(i, "AND")
            i += 2
        db.delete(f"[{TABLE_PREFIX}{dbtable}]", where_condition, top=1)
    qlik.reload_app(appid)

# db insert-update-delete öncesi değerlerde ince ayar yapılır
# tarih veya tarihsaat türünde alansa datestring belirtilen formatla parse edilir
# eğer alan string türünde ve boş stringse (""), null haline getirilir
def adjust_fields(row, table_fields):
    date_fields = [tf for tf in table_fields if tf["fieldtype"] == FieldType.DATE.value or tf["fieldtype"] == FieldType.DATETIME.value]
    str_fields = [tf for tf in table_fields if tf["fieldtype"] == FieldType.STRING.value]
    dfs = {df["fieldname"]: df["fieldformat"] for df in date_fields}
    ret_dict = {}
    for key, value in row.items():
        if key in dfs.keys():
            if value == '':
                ret_dict[key] = None
            else:
                ret_dict[key] = datetime.strptime(value, format_py_replace(dfs[key]))
        elif key in [sf["fieldname"] for sf in str_fields] and value is not None and value.strip() == '':
            ret_dict[key] = None
        else:
            ret_dict[key] = value
    return ret_dict

# ön-yüz addmeta modunda açıldıktan ve BI dev tarafından kayıt girilip Tümünü Kaydet'e basıldığında bu fonksiyona gelinir
# BI dev tarafından birden fazla qlik raporuna birden fazla tablo eklenebilir
# önce app'lere göre tablolar gruplanır, sonra tablolara göre field'lar gruplanır
# sonra field tanımları tek tek elden geçirilir, field'ta belirtilen appid-dbtable çiftine sahip diğer field tanımları da göz önünde bulundurulur
# bu işlemlerde en maliyetli kısım BI dev tarafından belirtilen qlik app'inin reload süresidir (bazı raporlarda 6-7 dk sürüyor)
# o yüzden reload işleminin tüm yeni field'ların tanımlanması bittikten sonra bir kere yapılması sağlanır
# field tanımları metadata tablosuna (QlikDataEntryMeta) da kaydedilir, meta qlik app'i de (Qlik Data Entry Configuration) reload edilir.
def add_meta_bulk(data):
    app_groups = {}
    table_groups = {}
    for row in data:
        if row["appid"] in app_groups.keys():
            if row["dbtable"] not in app_groups[row["appid"]]:
                app_groups[row["appid"]].append(row["dbtable"])
        else:
            app_groups[row["appid"]] = [row["dbtable"]]
        if row["dbtable"] in table_groups.keys():
            table_groups[row["dbtable"]].append(row)
        else:
            table_groups[row["dbtable"]] = [row]
    print("app_groups", app_groups)
    # print("table_groups", table_groups)
    for app in app_groups.keys():
        print(f"app guid: {app}")
        
        for table in app_groups[app]:
            app_tables = db.select(DB_CONN_META, ["*"], [("appid", "=", app)])
            print(f"db table: {table}")
            new_fields = [(
                row["fieldname"], 
                row["fieldheader"], 
                row["fieldtype"], 
                [["'" + pair.split(",")[0] + "'", "'" + pair.split(",")[1] + "'"] for pair in row["fieldenum"].split(";")] if row["fieldtype"] == FieldType.MENU.value else []
            ) for row in table_groups[table]]

            if len(app_tables) > 0:
                print("found existed tables for current app")
                table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", app), "AND", ("dbtable", "=", table)])
                if len(table_fields) > 0:
                    print("found existed fields for current table")
                    for row in table_groups[table]:
                        field_type = db_field_type(row["fieldtype"])
                        print("adding new column", row["fieldname"], "type:", row["fieldtype"])
                        db.table_add_column(f"[{TABLE_PREFIX}{table}]", f"[{row['fieldname']}] {field_type}")
                    other_tables = list(set([t["dbtable"] for t in app_tables if t["dbtable"] != table]))
                    other_tables_w_field = []
                    for t in other_tables:
                        other_table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", app), "AND", ("dbtable", "=", t)])
                        other_tables_w_field.append((t, [f["fieldname"] for f in other_table_fields]))
                    script = qlik.get_script(app)
                    new_field_names = [row["fieldname"] for row in table_groups[table]]
                    new_section = create_load_script(other_tables_w_field + [(table, [f["fieldname"] for f in table_fields] + new_field_names)])
                    new_script = modify_load_script(script, new_section)
                    qlik.set_script(app, new_script)
                    # qlik.reload_app(app)
                    sheet_ids = list(set([f["qsheetid"] for f in table_fields]))
                    table_qids = list(set([f["qtableid"] for f in table_fields]))
                    if len(sheet_ids) > 1:
                        raise LookupError("HATA: Aynı tablo ismiyle ilişkili birden fazla Qlik sheet id bulunuyor.")
                    if len(table_qids) > 1:
                        raise LookupError("HATA: Aynı tablo ismiyle ilişkili birden fazla Qlik table id bulunuyor.")            
                    curr_fields = [(
                        f["fieldname"], 
                        f["fieldheader"], 
                        f["fieldtype"], 
                        [["'" + pair.split(",")[0] + "'", "'" + pair.split(",")[1] + "'"] for pair in f["fieldenum"].split(";")] if f["fieldtype"] == FieldType.MENU.value else []
                    ) for f in table_fields]
                    custom_sheet_update(app, table_qids[0], curr_fields + new_fields)
                    for row in table_groups[table]:
                        row["qsheetid"], row["qtableid"] = sheet_ids[0], table_qids[0]
                else:
                    print("no existed fields for current table")
                    fields = [f"[{row['fieldname']}] {db_field_type(row['fieldtype'])}" for row in table_groups[table]]
                    db.create_table(f"[{TABLE_PREFIX}{table}]", fields)
                    other_tables = list(set([t["dbtable"] for t in app_tables]))
                    other_tables_w_field = []
                    for t in other_tables:
                        other_table_fields = db.select(DB_CONN_META, ["*"], [("appid", "=", app), "AND", ("dbtable", "=", t)])
                        other_tables_w_field.append((t, [f["fieldname"] for f in other_table_fields]))
                    script = qlik.get_script(app)                
                    new_field_names = [row["fieldname"] for row in table_groups[table]]
                    new_section = create_load_script(other_tables_w_field + [(table, new_field_names)])
                    new_script = modify_load_script(script, new_section)
                    qlik.set_script(app, new_script)
                    # qlik.reload_app(app)
                    sheet_id, table_qid = custom_sheet_create(app, table, new_fields)
                    for row in table_groups[table]:
                        row["qsheetid"], row["qtableid"] = sheet_id, table_qid
            else:
                print("no table found for current app")
                fields = [f"[{row['fieldname']}] {db_field_type(row['fieldtype'])}" for row in table_groups[table]]
                db.create_table(f"[{TABLE_PREFIX}{table}]", fields)
                script = qlik.get_script(app)
                new_field_names = [row["fieldname"] for row in table_groups[table]]
                new_script = create_load_script([(table, new_field_names)])
                qlik.set_script(app, script + new_script)
                # qlik.reload_app(app)            
                sheet_id, table_qid = custom_sheet_create(app, table, new_fields)
                for row in table_groups[table]:
                    row["qsheetid"], row["qtableid"] = sheet_id, table_qid

            for row in table_groups[table]:
                db.insert(DB_CONN_META, row.keys(), row.values())
        
        qlik.reload_app(app)
    
    qlik.reload_app(QLIK_META_APP)

# hiçbir yerde kullanılmıyor
def edit_meta_bulk(data):
    for row in data:
        if row["oldvalue"]["fieldname"] != row["newvalue"]["fieldname"]:
            where_condition = [(key, "=", value) if value is not None else f"{key} IS NULL" for key, value in row["oldvalue"].items()]
            i = 1
            while i < len(where_condition):
                where_condition.insert(i, "AND")
                i += 2
            set_condition = [(key, value) for key, value in row["newvalue"].items()]
            # db.update(DB_CONN_META, where_condition, set_condition)

# son kullanıcının veri girişi yaptığı (QDE_ prefix'li) tablodaki veriler getirilir
def get_rows(dbtable):
    return db.select(f"[{TABLE_PREFIX}{dbtable}]", ["*"])

# BI dev'in field tanımı yaptığı tablodaki (QlikDataEntryMeta) veriler getirilir
def get_rows_meta():
    return db.select(DB_CONN_META, ["*"])

# son kullanıcının yüklediği csv dosyası parse edilir ve iki boyutlu array olarak döndürülür
def upload_csv(file, name):
    file_byte = file.read()
    file_str = file_byte.decode("utf-8").split("\n")
    reader = csv.reader(file_str)
    return {"name": name, "rows": [r for r in reader if r != []]}

# son kullanıcının yüklediği xlsx dosyası parse edilir ve iki boyutlu array olarak döndürülür
def upload_xls(file, name):
    xls_wb = openpyxl.open(file)
    return {"name": name, "rows": [[c.value for c in r] for r in xls_wb.active.iter_rows()]}

# qlik raporundaki Data Load Script'e eklenecek section'ın yaratıldığı fonksiyondur
# bu script section'ı veri girişi yapılacak tablo ile (QDE_ prefix'li) qlik raporunun bağlandığı script'tir
# bir raporla ilişkili birden fazla tablo olsa bile tek section'da bağlantıları kurulur
def create_load_script(tables):
    new_line = '\r\n'
    return f"""{QLIK_TAB_PREFIX}{QLIK_SECTION}{new_line}
LIB CONNECT TO 'idmvmssqlp2 (idmd01_izvsktg)';
LET vDataEntryMode = 0;
{new_line.join([f'''
[{t[0]}]:
LOAD 
{(',' + new_line).join(t[1])};
SQL 
SELECT * FROM {db.DB_CONN_DB}.{db.DB_CONN_SCHEMA}.[{TABLE_PREFIX}{t[0]}];
''' for t in tables])}  
"""

# QDE_ prefix'li bir tablo ile ilişkisi olan bir rapora yeni bir alan eklendiğinde mevcut section ayrıştırılır
# mevcut section yeni script tarafından ezilir ve güncelleme sağlanmış olur
def modify_load_script(old_script, new_section):
    script_parts = old_script.split(f"{QLIK_TAB_PREFIX}{QLIK_SECTION}\r\n")
    if len(script_parts) == 1:
        raise NameError("HATA: " + QLIK_SECTION + " isimli section bulunamadı, ismi değişmiş veya tamamen silinmiş olabilir.")
    elif len(script_parts) > 2:
        raise NameError("HATA: Birden fazla sayıda " + QLIK_SECTION + " isimli section bulunuyor.")
    script_parts_2 = script_parts[1].split(QLIK_TAB_PREFIX)
    remained = QLIK_TAB_PREFIX.join([new_section] + script_parts_2[1:]) if len(script_parts_2) > 1 else new_section
    return script_parts[0] + remained
